import PySimpleGUI as sg      
from openpyxl import load_workbook
# PySimpleGUI
#has Magna Logo

sg.theme('DarkBlack')
layout = [
    [sg.Text('Enter Roll & Pitch Values')],
    [sg.Text('Roll', size =(15, 1)), sg.InputText()],
    [sg.Text('Pitch', size =(15, 1)), sg.InputText()],
    [sg.Submit(), sg.Cancel()]
]   
window = sg.Window('TestRun1', layout, icon='c:/Users/shoukulk/Desktop/work/Logos/MagnaLogo.ico')   
try:
    from ctypes import windll
    windll.shcore.SetProcessDpiAwareness(1)   
finally:      
 while True:
     
    event, values = window.read()   # type: ignore
    
    if event == 'Cancel':
        window.close()
    else:
        if(values[0].isnumeric() and values[1].isnumeric()):#type:ignore
            a = int(float(values[0]))
            b = int(float(values[1]))  
            if(a > 20 or b > 20):
                sg.popup('Enter Values Below 20', icon='c:/Users/shoukulk/Desktop/work/Logos/MagnaLogo.ico')
            else:
                sg.popup('Values Noted Successfully', icon='c:/Users/shoukulk/Desktop/work/Logos/MagnaLogo.ico')
                wb = load_workbook('C:/Users/shoukulk/Desktop/TestRunValues.xlsx')
                ws = wb.active   
                ws.cell(row =5, column =5).value = a # type: ignore
                ws.cell(row =5, column =6).value = b # type: ignore
                wb.save('C:/Users/shoukulk/Desktop/TestRunValues.xlsx')                         
        else:
            sg.popup('Only Integer Inputs are Accepted', icon='c:/Users/shoukulk/Desktop/work/Logos/MagnaLogo.ico')